import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 创建一个简单的 DataFrame
data = {'A': [1, 2, 3], 'B': ['A', 'B', 'C']}
df = pd.DataFrame(data)

# 将 DataFrame 保存为 Excel 文件
df.to_excel('example.xlsx', index=False)

# 加载 Excel 文件并选择工作表
wb = load_workbook('example.xlsx')
ws = wb.active

# 合并单元格: 假设我们要合并 A1:B1
ws.merge_cells('A1:B1')

# 居中对齐
alignment = Alignment(horizontal='center', vertical='center')

# 设置 A1 单元格的对齐方式为居中
ws['A1'].alignment = alignment

# 保存修改后的 Excel 文件
wb.save('example_modified.xlsx')
